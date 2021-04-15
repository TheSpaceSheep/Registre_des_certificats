from openpyxl import Workbook
from openpyxl.styles import *
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
import os

from registre_manager import Registre
from widgets import dialog
import language_selector as ls

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

thick_border = Border(left=Side(style='thick'),
                      right=Side(style='thick'),
                      top=Side(style='thick'),
                      bottom=Side(style='thick'))

thick_border_left = Border(left=Side(style='thick'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

thick_border_right = Border(left=Side(style='thin'),
                           right=Side(style='thick'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

thick_border_bottom = Border(left=Side(style='thin'),
                      right=Side(style='thin'),
                      top=Side(style='thin'),
                      bottom=Side(style='thick'))

thick_border_left_bottom = Border(left=Side(style='thick'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thick'))

thick_border_right_bottom = Border(left=Side(style='thin'),
                           right=Side(style='thick'),
                           top=Side(style='thin'),
                           bottom=Side(style='thick'))


def cm_to_points(cm):
    """converts centimeters into the weird excel unit : the points
    1 point = 1/72 of an inch... yeah..."""
    return cm*0.393701*72.

def inches_to_points(inches):
    return inches*72

def inches_to_cm(inches):
    return inches*2.54

MIN_ROW_HEIGHT = inches_to_points(0.18)
MAX_ROW_HEIGHT = cm_to_points(1)
MIN_COLUMN_WIDTH = inches_to_points(0.18)
MAX_COLUMN_WIDTH = cm_to_points(1)

MARGINS = 0.3  # inches
OFFSET = inches_to_points(0.1)

def generer_registre(registre, file="registre_des_certificats.xls"):
    """
    Generates a complete excel file with several sheets that represents the register
    The pages can be printed and physically displayed at relevant locations in the school.
    """
    wb = Workbook()

    # main page
    wb.active.title = ls.strings.REGISTER

    # names cannot be longer than 15 characters
    wb.active.column_dimensions['A'].width = 15

    wb.active.row_dimensions[1].height = cm_to_points(0.5)
    wb.active.row_dimensions[2].height = cm_to_points(4)
    wb.active.page_margins = PageMargins(MARGINS, MARGINS, MARGINS, MARGINS)  # most printers should handle these margins

    # first two rows
    j = 2
    cell = wb.active.cell(2, 1)
    cell.value = ls.strings.REGISTER_OF_CERTIFICATES
    cell.alignment = Alignment(wrap_text=True,
                               horizontal="center",
                               vertical="center")
    cell.border = thin_border
    cell.font = Font(name='arial', size=15, underline='single')

    for cat in registre.categories:
        cell = wb.active.cell(1, j)
        cell.value = cat
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_border
        first = True
        for c in registre.get_certificats(cat):
            cell = wb.active.cell(2, j)
            cell.value = c.nom
            cell.alignment = Alignment(horizontal="center",
                                       vertical="bottom",
                                       textRotation=90)
            cell.font = Font(name='arial', size=10)
            cell.border = thick_border_left if first else thin_border
            if first: first = False
            wb.active.column_dimensions[get_column_letter(j)].width = 3
            j += 1
        cell.border = thick_border_right
        wb.active.merge_cells(start_row=1,
                              start_column=j-len(registre.get_certificats(cat)),
                              end_row=1,
                              end_column=j-1)

    """
    ValueError: Value must be one of {'lightDown', 'darkGray', 'lightHorizontal',
    'lightVertical', 'darkDown', 'gray0625', 'solid', 'gray125', 'lightGray', 'darkGrid',
    'lightGrid', 'lightUp', 'mediumGray', 'darkTrellis', 'darkVertical', 'lightTrellis', 'darkUp',
    'darkHorizontal'

    """
    # member rows
    i = 3
    k = 0
    for m in registre.membres:
        j = 2
        header_height = wb.active.row_dimensions[1].height + wb.active.row_dimensions[2].height
        height_hint = (cm_to_points(29.7) - header_height - inches_to_points(2*MARGINS) - OFFSET) / len(registre.membres)
        if height_hint < MIN_ROW_HEIGHT and height_hint > MIN_ROW_HEIGHT/2:
            wb.active.row_dimensions[i].height = 2*height_hint
        elif height_hint > MAX_ROW_HEIGHT:
            wb.active.row_dimensions[i].height = MAX_ROW_HEIGHT
        else:
            wb.active.row_dimensions[i].height = height_hint
        cell = wb.active.cell(i, 1)
        cell.value = m.id
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_border_right_bottom if i%4==2 else thick_border_right
        for cat in registre.categories:
            first = True
            for cert in registre.get_certificats(cat):
                cell = wb.active.cell(i, j)
                if registre.registre[m, cert] == Registre.NonCertifie:
                    cell.fill = PatternFill(bgColor="000000", fill_type = None)
                    k+=1
                elif registre.registre[m, cert] == Registre.Certifie:
                    cell.fill = PatternFill(bgColor="00FF00", fill_type ="gray0625")
                elif registre.registre[m, cert] == Registre.Certificateur:
                    cell.fill = PatternFill(bgColor="0000FF", fill_type ="gray0625")
                elif registre.registre[m, cert] == Registre.CertificatPerdu:
                    cell.fill = PatternFill(bgColor="888888", fill_type ="gray0625")

                cell.alignment = Alignment(horizontal="center", vertical="center")
                if first:
                    cell.border = thick_border_left_bottom if i%4==2 else thick_border_left
                else:
                    cell.border = thick_border_bottom if i%4==2 else thin_border
                first = False
                j += 1
            cell.border = thick_border_right_bottom if i%4==2 else thick_border_right
        i += 1

    # page for each category 
    for cat in registre.categories:
        wb.create_sheet(cat)
        wb.active = wb[cat]

        wb.active.column_dimensions['A'].width = 15
        wb.active.row_dimensions[1].height = cm_to_points(0.5)
        wb.active.row_dimensions[2].height = cm_to_points(4)
        wb.active.page_margins = PageMargins(MARGINS, MARGINS, MARGINS, MARGINS)  # most printers should handle these margins

        # first two rows
        cell = wb.active.cell(2, 1)
        cell.value = f"Certificats {cat}"
        cell.alignment = Alignment(wrap_text=True,
                                   horizontal="center",
                                   vertical="center")
        cell.font = Font(name='arial', size=15, underline='single')
        cell.border = thin_border

        cell = wb.active.cell(1, 2)
        cell.value = f"{cat}"
        cell.alignment = Alignment(horizontal="center",
                                   vertical="center")
        cell.border = thick_border

        wb.active.merge_cells(start_row=1,
                              start_column=2,
                              end_row=1,
                              end_column=len(registre.get_certificats(cat))+1)
        j = 2
        first = True
        for cert in registre.get_certificats(cat):
            cell = wb.active.cell(2, j)
            cell.value = cert.nom
            cell.alignment = Alignment(horizontal="center",
                                       vertical="bottom",
                                       textRotation=90)
            cell.font = Font(name='arial', size=10)
            cell.border = thick_border_left if first else thin_border
            first = False
            j += 1
        cell.border = thick_border_right

        i = 3
        # member rows
        for m in registre.membres:
            header_height = wb.active.row_dimensions[1].height + wb.active.row_dimensions[2].height
            height_hint = (cm_to_points(29.7) - header_height - inches_to_points(2*MARGINS) - OFFSET) / len(registre.membres)
            if height_hint < MIN_ROW_HEIGHT and height_hint > MIN_ROW_HEIGHT/2:
                wb.active.row_dimensions[i].height = 2*height_hint
            elif height_hint > MAX_ROW_HEIGHT:
                wb.active.row_dimensions[i].height = MAX_ROW_HEIGHT
            else:
                wb.active.row_dimensions[i].height = height_hint
            cell = wb.active.cell(i, 1)
            cell.value = m.id
            cell.alignment = Alignment(horizontal="center",
                                       vertical="center")
            cell.border = thick_border_right_bottom if i%4==2 else thick_border_right
            j = 2
            first = True
            for cert in registre.get_certificats(cat):
                cell = wb.active.cell(i, j)
                if registre.registre[m, cert] == Registre.NonCertifie:
                    cell.fill = PatternFill(bgColor="FFFFFF", fill_type = None)
                elif registre.registre[m, cert] == Registre.Certifie:
                    cell.fill = PatternFill(bgColor="00FF00", fill_type ="gray0625")
                    cell.value = "c"
                elif registre.registre[m, cert] == Registre.Certificateur:
                    cell.fill = PatternFill(bgColor="0000FF", fill_type ="gray0625")
                    cell.value = "C"
                elif registre.registre[m, cert] == Registre.CertificatPerdu:
                    cell.fill = PatternFill(bgColor="888888", fill_type ="gray0625")
                    cell.value = "P"
                cell.alignment = Alignment(horizontal="center",
                                           vertical="center")
                if first:
                    cell.border = thick_border_left_bottom if i%4==2 else thick_border_left
                else:
                    cell.border = thick_border_bottom if i%4==2 else thin_border
                first = False
                wb.active.column_dimensions[get_column_letter(j)].width = 5
                j += 1
            cell.border = thick_border_right_bottom if i%4==2 else thick_border_right
            i += 1

    wb.save(file)

    # lauching libreoffice
    # on windows
    if os.name == "nt":
        import winreg
        try:
            value = winreg.QueryValue(winreg.HKEY_LOCAL_MACHINE, "SOFTWARE\\LibreOffice\\UNO\\InstallPath")
            os.system(f"start /B \"{value}\\soffice\" {file}")
        except FileNotFoundError:
            try:
                value = winreg.QueryValue(winreg.HKEY_LOCAL_MACHINE, "SOFTWARE\\Wow6432Node\\LibreOffice\\UNO\\InstallPath")
                os.system(f"start /B \"{value}\\soffice\" {file}")
            except FileNotFoundError:
                dialog(ls.strings.CANT_OPEN_LIBREOFFICE)
                os.system(f"explorer {os.getcwd()}/printable")
    # on linux
    elif os.name == "posix":
        os.system(f"libreoffice {file}&")


def center_all_cells(wb):
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

